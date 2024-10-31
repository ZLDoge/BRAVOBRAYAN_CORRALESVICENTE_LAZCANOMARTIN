from django.shortcuts import render, redirect
from .forms import archivoform, Intento1Form

# from .utils import comparar_registros, generar_informe 

# Create your views here.
def home(request):
    return render(request, 'home.html')

# def form_archivo(request):
#     success_message = None  # Variable para el mensaje de éxito
    
#     if request.method == 'POST':
#         form = Intento1Form(request.POST, request.FILES)
#         if form.is_valid():
#             form.save()
#             success_message = "Archivo y correo subidos correctamente."  # Mensaje de éxito
#             return redirect('prueba')
#         else:
#             # Esto asegura que incluso si el formulario no es válido, devuelva una respuesta
#             return render(request, 'form_archivo.html', {'form': form})
#     else:
#         form = Intento1Form()  # Mostrar formulario vacío si la solicitud es GET
    
#     # Siempre se devuelve una respuesta, ya sea para GET o POST
#     return render(request, 'form_archivo.html', {'form': form, 'success_message': success_message})

from django.contrib.auth import authenticate, login
from django.contrib import messages

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)  # Iniciar sesión del usuario (puede ser superusuario)
            return redirect('form_archivo')  # Redirige al área protegida
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    
    return render(request, 'intento.html')




def prueba(request):
    return render(request, 'prueba.html')

def intento(request):
    return render(request, 'intento.html')

###############CALCULO######################
from django.shortcuts import render
from django.http import HttpResponse
from .forms import Intento1Form
from .models import bbdd
import pandas as pd
from io import BytesIO
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt

def form_archivo(request):
    if request.method == 'POST':
        form = Intento1Form(request.POST, request.FILES)
        if form.is_valid():
            registro = form.save()
            archivo1 = registro.documento1.path
            archivo2 = registro.documento2.path

            # Leer los archivos como DataFrames de Pandas
            df_clientes = pd.read_csv(archivo1)
            df_base_datos = pd.read_csv(archivo2)

            # Procesar y comparar registros
            resultados = comparar_registros(df_clientes, df_base_datos)
            df_clientes = resultados[0]
            resumen = resultados[1:]


            # Generar el informe en memoria
            informe = generar_informe(df_clientes, *resumen)

            # Enviar el archivo Excel generado al usuario
            response = HttpResponse(
                informe, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=informe_puenteo.xlsx'
            return response  # Descargar el informe como respuesta
    else:
        form = Intento1Form()

    return render(request, 'form_archivo.html', {'form': form})

# Función para comparar registros
def comparar_registros(clientes, base_datos):
    clientes['Similitud'] = 0
    clientes['Encontrado'] = 'No'
    clientes['Registro_BBDD'] = ''
    clientes['Sugerencia'] = ''

    for index, cliente in clientes.iterrows():
        nombre_apellido_cliente = f"{cliente['Nombre']} {cliente['Apellido']}"
        mejor_coincidencia = process.extractOne(
            nombre_apellido_cliente,
            base_datos['Nombre'] + ' ' + base_datos['Apellido'],
            scorer=fuzz.token_sort_ratio
        )

        if mejor_coincidencia and mejor_coincidencia[1] >= 80:
            clientes.at[index, 'Similitud'] = mejor_coincidencia[1]
            clientes.at[index, 'Encontrado'] = 'Sí'
            especialidad = base_datos.loc[
                (base_datos['Nombre'] + ' ' + base_datos['Apellido'] == mejor_coincidencia[0]), 'Especialidad'
            ].values
            clientes.at[index, 'Registro_BBDD'] = especialidad[0] if len(especialidad) > 0 else ''
        elif 65 <= mejor_coincidencia[1] < 80:
            clientes.at[index, 'Similitud'] = mejor_coincidencia[1]
            clientes.at[index, 'Sugerencia'] = 'Revisar posible coincidencia'

    total_registros = clientes.shape[0]
    total_encontrados = (clientes['Encontrado'] == 'Sí').sum()
    total_no_encontrados = total_registros - total_encontrados

    porcentaje_encontrados = (total_encontrados / total_registros) * 100
    porcentaje_no_encontrados = 100 - porcentaje_encontrados

    return (clientes, total_registros, total_encontrados, total_no_encontrados,
            round(porcentaje_encontrados, 2), round(porcentaje_no_encontrados, 2))

# Función para generar el informe con gráfico
def generar_informe(clientes, total_registros, total_encontrados, total_no_encontrados, porcentaje_encontrados, porcentaje_no_encontrados):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"

    # Configurar título y encabezados
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Puenteo Realizado - IQVIA'
    ws['A1'].font = Font(size=20, bold=True, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    fill_color_titulo = PatternFill(start_color='0C7BC0', end_color='0C7BC0', fill_type='solid')
    for row in ws['A1:D1']:
        for cell in row:
            cell.fill = fill_color_titulo

    # Encabezados
    encabezados = ['Nombre', 'Apellido', 'Similitud', 'Encontrado', 'Registro_BBDD', 'Sugerencia']
    fill_color = PatternFill(start_color='0C7BC0', end_color='0C7BC0', fill_type='solid')

    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=3, column=col_num, value=encabezado)
        celda.fill = fill_color
        celda.font = Font(bold=True, color='FFFFFF')
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for r in dataframe_to_rows(clientes[['Nombre', 'Apellido', 'Similitud', 'Encontrado', 'Registro_BBDD', 'Sugerencia']], index=False, header=False):
        ws.append(r)

    # Agregar resumen al final
    ws.append([])  
    ws.append(['Total Registros Procesados', total_registros])
    ws.append(['Total Registros Encontrados', total_encontrados])
    ws.append(['Total Registros No Encontrados', total_no_encontrados])
    ws.append(['Porcentaje Encontrados', f"{porcentaje_encontrados}%"])
    ws.append(['Porcentaje No Encontrados', f"{porcentaje_no_encontrados}%"])

    # Crear el gráfico de torta
    etiquetas = ['Encontrados', 'No Encontrados']
    datos = [porcentaje_encontrados, porcentaje_no_encontrados]
    colores = ['#0C7BC0', '#87CEEB']

    plt.figure(figsize=(5, 5))
    plt.pie(datos, labels=etiquetas, autopct='%1.1f%%', startangle=90, colors=colores, wedgeprops={'edgecolor': 'black'})
    plt.title('Distribución de registros encontrados vs no encontrados')

    nombre_grafico = 'grafico_torta.png'
    plt.savefig(nombre_grafico)
    plt.close()

    # Insertar la imagen en el Excel
    img = Image(nombre_grafico)
    ws.add_image(img, 'F1')

    wb.save(output)
    output.seek(0)
    return output
