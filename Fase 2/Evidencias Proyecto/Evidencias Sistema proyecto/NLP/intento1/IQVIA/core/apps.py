from django.apps import AppConfig


class CoreConfig(AppConfig):
    name = 'core'


from flask import Flask, request, render_template, send_file
import pandas as pd
from fuzzywuzzy import process
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from io import BytesIO
from tqdm import tqdm

app = Flask(__name__)

# Ruta para la página principal
@app.route('/')
def index():
    return render_template('form_archivo.html')

# Ruta para procesar los archivos y generar el informe
@app.route('/prueba', methods=['POST'])
def procesar():
    # Cargar los archivos subidos
    clientes_file = request.files['clientes']
    base_datos_file = request.files['base_datos']

    # Leer archivos CSV a DataFrames de Pandas
    clientes = pd.read_csv(clientes_file)
    base_datos = pd.read_csv(base_datos_file)

    # Llamar a la función de comparación de registros
    clientes, resumen = comparar_registros(clientes, base_datos)

    # Generar el informe en memoria
    informe = generar_informe_en_memoria(clientes, resumen)

    # Enviar el archivo Excel generado al usuario para su descarga
    return send_file(informe, download_name='informe_corporativo.xlsx', as_attachment=True)

# Función para comparar registros entre dos DataFrames
def comparar_registros(clientes, base_datos):
    clientes['Similitud'] = 0
    clientes['Encontrado'] = 'No'
    clientes['Registro_BBDD'] = ''
    clientes['Sugerencia'] = ''

    for index, cliente in tqdm(clientes.iterrows(), total=clientes.shape[0]):
        nombre_completo_cliente = f"{cliente['Nombre']} {cliente['Apellido']}"
        mejor_coincidencia = process.extractOne(
            nombre_completo_cliente,
            base_datos['Nombre'] + ' ' + base_datos['Apellido'],
            scorer=fuzz.token_sort_ratio
        )

        if mejor_coincidencia and mejor_coincidencia[1] >= 80:
            clientes.at[index, 'Similitud'] = mejor_coincidencia[1]
            clientes.at[index, 'Encontrado'] = 'Sí'
            especialidad = base_datos.loc[
                (base_datos['Nombre'] + ' ' + base_datos['Apellido'] == mejor_coincidencia[0]), 'Especialidad'
            ].values
            clientes.at[index, 'Registro_BBDD'] = especialidad[0] if len(especialidad) > 0 else ''
        elif 65 <= mejor_coincidencia[1] < 80:
            clientes.at[index, 'Similitud'] = mejor_coincidencia[1]
            clientes.at[index, 'Sugerencia'] = 'Revisar posible coincidencia'

    # Resumen de resultados
    resumen = {
        'total_registros': clientes.shape[0],
        'total_encontrados': (clientes['Encontrado'] == 'Sí').sum(),
        'total_no_encontrados': (clientes['Encontrado'] == 'No').sum(),
    }
    resumen['porcentaje_encontrados'] = (resumen['total_encontrados'] / resumen['total_registros']) * 100
    resumen['porcentaje_no_encontrados'] = 100 - resumen['porcentaje_encontrados']
    return clientes, resumen

# Función para generar el informe Excel en memoria
def generar_informe_en_memoria(clientes, resumen):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"

    # Título
    ws['A1'] = 'Informe Comparativo'
    ws.merge_cells('A1:F1')
    ws['A1'].style = 'Headline 1'

    # Encabezados
    encabezados = ['Nombre', 'Apellido', 'Similitud', 'Encontrado', 'Registro_BBDD', 'Sugerencia']
    ws.append(encabezados)

    # Datos de clientes
    for _, row in clientes.iterrows():
        ws.append([row['Nombre'], row['Apellido'], row['Similitud'], row['Encontrado'], row['Registro_BBDD'], row['Sugerencia']])

    # Agregar resumen
    ws.append([])
    ws.append(['Resumen', ''])
    ws.append(['Total Registros', resumen['total_registros']])
    ws.append(['Total Encontrados', resumen['total_encontrados']])
    ws.append(['Total No Encontrados', resumen['total_no_encontrados']])
    ws.append(['Porcentaje Encontrados', f"{resumen['porcentaje_encontrados']}%"])
    ws.append(['Porcentaje No Encontrados', f"{resumen['porcentaje_no_encontrados']}%"])

    wb.save(output)
    output.seek(0)
    return output

if __name__ == '__main__':
    app.run(debug=True)
